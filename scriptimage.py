import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import openpyxl
import re

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Fonction pour prétraiter l'image
def preprocess_image(image_path):
    image = Image.open(image_path)
    image = image.convert('L')  # Convertir en niveaux de gris
    image = image.filter(ImageFilter.MedianFilter())  # Filtre pour désencombrer l'image
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(2)  # Améliorer le contraste
    image = image.point(lambda x: 0 if x < 128 else 255, '1')  # Binarisation
    return image  # Retourner l'image traitée

# Fonction pour extraire le texte d'une image
def extract_text_from_image(image_path):
    image = preprocess_image(image_path)
    # Ajuster les paramètres de Tesseract pour des textes plus complexes
    custom_config = r'--oem 3 --psm 6'
    text = pytesseract.image_to_string(image, config=custom_config)
    return text

# Fonction pour extraire le nom et l'application du texte
def extract_info_from_text(text):
    nom = re.search(r'Nom:\s*(.*)', text)
    application = re.search(r'Application\s*:\s*(.*)', text)
    nom_value = nom.group(1).strip() if nom else 'N/A'
    application_value = application.group(1).strip() if application else 'N/A'
    return nom_value, application_value

# Fonction pour écrire le texte dans un fichier Excel
def write_text_to_excel(nom, application, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Texte extrait"
    ws.cell(row=1, column=1, value="Nom")
    ws.cell(row=1, column=2, value="Application")
    ws.cell(row=2, column=1, value=nom)
    ws.cell(row=2, column=2, value=application)
    wb.save(excel_path)

# Chemin de l'image à traiter
image_path = 'C:\\Users\\Aya\\Desktop\\for.png'

# Chemin du fichier Excel de sortie
excel_path = r'C:\Users\Aya\Desktop\Classeur2.xlsx'

# Extraire le texte de l'image
text = extract_text_from_image(image_path)

# Extraire les informations du texte
nom, application = extract_info_from_text(text)

# Écrire les informations dans le fichier Excel
write_text_to_excel(nom, application, excel_path)

print(f"Le texte a été extrait et sauvegardé dans {excel_path}")
