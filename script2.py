import pytesseract
from PIL import Image
import openpyxl
import re

# Chemin vers l'exécutable Tesseract (modifie-le si nécessaire)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Fonction pour extraire le texte d'une image
def extract_text_from_image(image_path):
    image = Image.open(image_path)
    text = pytesseract.image_to_string(image)
    return text

# Fonction pour extraire le nom et l'application du texte
def extract_info_from_text(text):
    nom = re.search(r'Nom:\s*(.*)', text)
    application = re.search(r'Application\s*:\s*(.*)', text)
    nom_value = nom.group(1) if nom else 'N/A'
    application_value = application.group(1) if application else 'N/A'
    return nom_value, application_value

# Fonction pour écrire le texte dans un fichier Excel
def write_text_to_excel(nom, application, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Texte extrait"
    

    
    # Écrire les valeurs
    ws.cell(row=1, column=1, value=nom)
    ws.cell(row=1, column=2, value=application)
    
    wb.save(excel_path)

# Chemin de l'image à traiter
image_path = 'C:\\Users\\hayfa\\Desktop\\for.jpg'

# Chemin du fichier Excel de sortie
excel_path = r'C:\Users\hayfa\Desktop\Classeur2.xlsx'

# Extraire le texte de l'image
text = extract_text_from_image(image_path)

# Extraire les informations du texte
nom, application = extract_info_from_text(text)

# Écrire les informations dans le fichier Excel
write_text_to_excel(nom, application, excel_path)

print(f"Le texte a été extrait et sauvegardé dans {excel_path}")
